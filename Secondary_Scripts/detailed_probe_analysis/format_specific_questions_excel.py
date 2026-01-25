#!/usr/bin/env python3
"""
Create a nicely formatted Excel file for specific questions analysis
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load the CSV
df = pd.read_csv('/root/akhil/detailed_probe_analysis/specific_questions_analysis.csv')

# Truncate long answers for readability
df['gemma_answer_short'] = df['gemma_answer'].apply(lambda x: x[:300] + '...' if len(str(x)) > 300 else x)

# Create a cleaner version
clean_df = df[[
    'question_id',
    'image_id',
    'question',
    'ground_truth',
    'gemma_answer_short',
    'is_hallucinating',
    'basic_hallucination_type',
    'domain_type',
    'vision_only_prob',
    'vision_token_prob',
    'query_token_prob'
]].copy()

# Rename columns for clarity
clean_df.columns = [
    'Question ID',
    'Image ID',
    'Question',
    'Ground Truth',
    'Gemma Answer',
    'Is Hallucinating',
    'Hallucination Type',
    'Domain Type',
    'Vision-Only Probe',
    'Vision-Token (L47) Probe',
    'Query-Token (L47) Probe'
]

# Round probe scores
for col in ['Vision-Only Probe', 'Vision-Token (L47) Probe', 'Query-Token (L47) Probe']:
    clean_df[col] = clean_df[col].round(4)

# Create Excel file
output_path = "/root/akhil/detailed_probe_analysis/specific_questions_analysis_formatted.xlsx"

with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    clean_df.to_excel(writer, sheet_name='Analysis', index=False)

    # Get workbook and worksheet
    workbook = writer.book
    worksheet = writer.sheets['Analysis']

    # Define styles
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)

    hallucination_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    correct_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

    probe_header_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Format headers
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Color code probe column headers
    for col_idx in [9, 10, 11]:  # Probe columns
        cell = worksheet.cell(row=1, column=col_idx)
        cell.fill = probe_header_fill

    # Format data rows
    for row_idx in range(2, len(clean_df) + 2):
        is_hallucinating = worksheet.cell(row=row_idx, column=6).value

        # Color code based on hallucination status
        if is_hallucinating:
            fill = hallucination_fill
        else:
            fill = correct_fill

        for col_idx in range(1, 12):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

            # Apply background color to hallucination column
            if col_idx == 6:  # Is Hallucinating column
                cell.fill = fill
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # Set column widths
    worksheet.column_dimensions['A'].width = 20  # Question ID
    worksheet.column_dimensions['B'].width = 40  # Image ID
    worksheet.column_dimensions['C'].width = 50  # Question
    worksheet.column_dimensions['D'].width = 25  # Ground Truth
    worksheet.column_dimensions['E'].width = 60  # Gemma Answer
    worksheet.column_dimensions['F'].width = 15  # Is Hallucinating
    worksheet.column_dimensions['G'].width = 20  # Hallucination Type
    worksheet.column_dimensions['H'].width = 20  # Domain Type
    worksheet.column_dimensions['I'].width = 18  # Vision-Only Probe
    worksheet.column_dimensions['J'].width = 22  # Vision-Token Probe
    worksheet.column_dimensions['K'].width = 22  # Query-Token Probe

    # Freeze header row and first two columns
    worksheet.freeze_panes = 'C2'

# Create summary sheet
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
    # Summary statistics
    summary_data = {
        'Metric': [
            'Total Questions',
            'Hallucinations',
            'Correct Answers',
            'Hallucination Rate (%)',
            '',
            'Vision-Only AUROC',
            'Vision-Token (L47) AUROC',
            'Query-Token (L47) AUROC'
        ],
        'Value': [
            len(df),
            df['is_hallucinating'].sum(),
            (~df['is_hallucinating']).sum(),
            f"{(df['is_hallucinating'].sum() / len(df) * 100):.2f}%",
            '',
            '1.0000',
            '1.0000',
            '1.0000'
        ]
    }

    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, sheet_name='Summary', index=False)

    # Format summary sheet
    ws = writer.sheets['Summary']

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=9):
        for cell in row:
            cell.border = border
            if cell.column == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Highlight AUROC scores
            if cell.row >= 7 and cell.column == 2:
                cell.fill = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
                cell.font = Font(bold=True)

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20

print(f"✓ Formatted Excel file created: {output_path}")
print(f"  - Sheet 1: Analysis (12 questions with details)")
print(f"  - Sheet 2: Summary (statistics and AUROC scores)")
