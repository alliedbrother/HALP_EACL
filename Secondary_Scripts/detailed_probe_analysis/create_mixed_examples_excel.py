#!/usr/bin/env python3
"""
Create a well-formatted Excel file with mixed hallucination examples
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Load the mixed hallucination images CSV
mixed_images_df = pd.read_csv('/root/akhil/detailed_probe_analysis/mixed_hallucination_images.csv')

# Load detailed metadata
metadata_df = pd.read_csv('/root/akhil/final_data/sampled_10k_with_hallucination_types.csv')

# Model CSV files
MODEL_CSVS = {
    'Gemma3-12B': '/root/akhil/FInal_CSV_Hallucination/gemma3_manually_reviewed.csv',
    'FastVLM-7B': '/root/akhil/FInal_CSV_Hallucination/fastvlm_manually_reviewed.csv',
    'LLaVA-Next-8B': '/root/akhil/FInal_CSV_Hallucination/llava_manually_reviewed.csv',
    'Molmo-V1': '/root/akhil/FInal_CSV_Hallucination/molmo_manually_reviewed.csv',
    'Qwen2.5-VL-7B': '/root/akhil/FInal_CSV_Hallucination/qwen25vl_manually_reviewed.csv',
    'SmolVLM2-2.2B': '/root/akhil/FInal_CSV_Hallucination/smolvlm_manually_reviewed.csv',
    'Llama-3.2-11B': '/root/akhil/FInal_CSV_Hallucination/llama32_manually_reviewed.csv',
    'Phi4-VL': '/root/akhil/FInal_CSV_Hallucination/phi4vl_manually_reviewed.csv',
}

# Create metadata lookup
metadata_lookup = {}
for _, row in metadata_df.iterrows():
    metadata_lookup[row['question_id']] = {
        'basic_hallucination_type': row.get('basic_hallucination_type', 'Unknown'),
        'domain_type': row.get('domain_type', 'Unknown'),
        'answer_type': row.get('answer_type', 'Unknown')
    }

print("Creating Excel with mixed hallucination examples...")

all_examples = []

for model_name, csv_path in MODEL_CSVS.items():
    print(f"Processing {model_name}...")

    # Load model CSV
    model_df = pd.read_csv(csv_path, low_memory=False)

    # Convert hallucination label
    model_df['is_hallucinating'] = model_df['is_hallucinating_manual'].apply(
        lambda x: True if (isinstance(x, str) and x.lower() in ['true', '1', 'yes']) or x == True or x == 1 else False
    )

    # Get mixed images for this model
    model_mixed = mixed_images_df[mixed_images_df['model'] == model_name]

    if len(model_mixed) == 0:
        continue

    # Add metadata to model_df
    model_df['basic_hallucination_type'] = model_df['question_id'].map(
        lambda qid: metadata_lookup.get(qid, {}).get('basic_hallucination_type', 'Unknown')
    )
    model_df['domain_type'] = model_df['question_id'].map(
        lambda qid: metadata_lookup.get(qid, {}).get('domain_type', 'Unknown')
    )

    # Track which types we've covered
    covered_basic_types = set()
    covered_domain_types = set()

    examples_found = 0

    # Iterate through mixed images (sorted by most balanced)
    for _, mixed_img in model_mixed.iterrows():
        if examples_found >= 12:  # Limit examples per model
            break

        image_id = mixed_img['image_id']
        image_questions = model_df[model_df['image_id'] == image_id]

        hall_questions = image_questions[image_questions['is_hallucinating'] == True]
        no_hall_questions = image_questions[image_questions['is_hallucinating'] == False]

        for _, hall_q in hall_questions.iterrows():
            basic_type = hall_q['basic_hallucination_type']
            domain_type = hall_q['domain_type']

            if basic_type not in covered_basic_types or domain_type not in covered_domain_types:
                if len(no_hall_questions) > 0:
                    no_hall_q = no_hall_questions.iloc[0]

                    # Add hallucinated question
                    all_examples.append({
                        'Model': model_name,
                        'Image': image_id,
                        'Total_Questions': len(image_questions),
                        'Hallucinated_Count': int(mixed_img['hallucinated_questions']),
                        'Correct_Count': int(mixed_img['no_hallucination_questions']),
                        'Question_Type': 'HALLUCINATED',
                        'Question_ID': hall_q['question_id'],
                        'Question': hall_q['question'],
                        'Ground_Truth': str(hall_q['ground_truth_answer']),
                        'Model_Answer': str(hall_q['model_answer']),
                        'Basic_Hallucination_Type': basic_type,
                        'Domain_Type': domain_type
                    })

                    # Add correct question
                    all_examples.append({
                        'Model': model_name,
                        'Image': image_id,
                        'Total_Questions': len(image_questions),
                        'Hallucinated_Count': int(mixed_img['hallucinated_questions']),
                        'Correct_Count': int(mixed_img['no_hallucination_questions']),
                        'Question_Type': 'CORRECT',
                        'Question_ID': no_hall_q['question_id'],
                        'Question': no_hall_q['question'],
                        'Ground_Truth': str(no_hall_q['ground_truth_answer']),
                        'Model_Answer': str(no_hall_q['model_answer']),
                        'Basic_Hallucination_Type': no_hall_q.get('basic_hallucination_type', 'N/A'),
                        'Domain_Type': no_hall_q.get('domain_type', 'N/A')
                    })

                    covered_basic_types.add(basic_type)
                    covered_domain_types.add(domain_type)
                    examples_found += 1
                    break

    print(f"  Found {examples_found} example pairs")

# Create DataFrame
df = pd.DataFrame(all_examples)

# Create Excel with formatting
output_path = "/root/akhil/detailed_probe_analysis/mixed_hallucination_examples.xlsx"

with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Mixed Examples', index=False)

    # Get workbook and worksheet
    workbook = writer.book
    worksheet = writer.sheets['Mixed Examples']

    # Define styles
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)

    hallucinated_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    correct_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Format headers
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Format data rows
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        question_type = worksheet[f'F{row_idx}'].value

        # Color code based on question type
        if question_type == 'HALLUCINATED':
            fill = hallucinated_fill
        else:
            fill = correct_fill

        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

            # Apply background color to key columns
            if cell.column in [6, 7, 8, 9, 10]:  # Question_Type through Model_Answer
                cell.fill = fill

    # Set column widths
    worksheet.column_dimensions['A'].width = 18  # Model
    worksheet.column_dimensions['B'].width = 20  # Image
    worksheet.column_dimensions['C'].width = 12  # Total_Questions
    worksheet.column_dimensions['D'].width = 15  # Hallucinated_Count
    worksheet.column_dimensions['E'].width = 15  # Correct_Count
    worksheet.column_dimensions['F'].width = 15  # Question_Type
    worksheet.column_dimensions['G'].width = 25  # Question_ID
    worksheet.column_dimensions['H'].width = 50  # Question
    worksheet.column_dimensions['I'].width = 25  # Ground_Truth
    worksheet.column_dimensions['J'].width = 60  # Model_Answer
    worksheet.column_dimensions['K'].width = 25  # Basic_Hallucination_Type
    worksheet.column_dimensions['L'].width = 25  # Domain_Type

    # Freeze header row and first two columns
    worksheet.freeze_panes = 'C2'

# Create summary sheet
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
    # Summary by model
    summary_data = []
    for model in df['Model'].unique():
        model_data = df[df['Model'] == model]
        hall_count = len(model_data[model_data['Question_Type'] == 'HALLUCINATED'])

        summary_data.append({
            'Model': model,
            'Example_Pairs': hall_count,
            'Basic_Hallucination_Types': len(model_data[model_data['Question_Type'] == 'HALLUCINATED']['Basic_Hallucination_Type'].unique()),
            'Domain_Types': len(model_data[model_data['Question_Type'] == 'HALLUCINATED']['Domain_Type'].unique())
        })

    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, sheet_name='Summary', index=False)

    # Format summary sheet
    ws = writer.sheets['Summary']
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 20

print(f"\n{'='*80}")
print(f"Excel file created: {output_path}")
print(f"Total examples: {len(df)} rows ({len(df)//2} pairs)")
print(f"Sheets: 'Mixed Examples' (main data), 'Summary' (overview)")
print('='*80)
